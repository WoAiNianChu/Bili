import os
import re
from datetime import datetime
import threading
from openpyxl import load_workbook

class ExcelComparator:
    """Excel文件对比核心类（完整版）"""

    def __init__(self):
        self.product_data = {}
        self.kitchen_data = {}
        threading.Thread(target=self.lazy_import_openpyxl).start()

    def lazy_import_openpyxl(self):
        """延迟加载openpyxl模块"""
        global load_workbook
        from openpyxl import load_workbook

    def auto_detect_files(self):
        """自动检测必要文件"""
        while True:
            os.system('cls' if os.name == 'nt' else 'clear')
            print("正在扫描目录...")
            current_dir = os.getcwd()
            files = {
                "product": {"pattern": "产品统计表", "found": None, "name": "产品统计表"},
                "kitchen": {"pattern": "厨房", "found": None, "name": "厨房用表"}
            }

            # 文件检测逻辑
            for filename in os.listdir(current_dir):
                filepath = os.path.join(current_dir, filename)
                if not os.path.isfile(filepath):
                    continue

                # 产品统计表匹配
                if files["product"]["pattern"] in filename and "_对比结果" not in filename:
                    if not files["product"]["found"] or \
                       os.path.getmtime(filepath) > os.path.getmtime(files["product"]["found"]):
                        files["product"]["found"] = filepath
                
                # 厨房用表匹配
                if files["kitchen"]["pattern"] in filename:
                    if not files["kitchen"]["found"] or \
                       os.path.getmtime(filepath) > os.path.getmtime(files["kitchen"]["found"]):
                        files["kitchen"]["found"] = filepath

            # 文件存在性检查
            missing_files = [f["name"] for f in files.values() if not f["found"]]
            if not missing_files:
                print("\n" + "="*50)
                print("检测到以下文件：")
                for f_type in files.values():
                    mtime = datetime.fromtimestamp(os.path.getmtime(f_type["found"]))
                    print(f"[{f_type['name']}]")
                    print(f"文件名：{os.path.basename(f_type['found'])}")
                    print(f"修改时间：{mtime.strftime('%Y-%m-%d %H:%M:%S')}\n")
                print("="*50)
                
                # 用户确认
                choice = input("是否开始对比？(Y/N): ").strip().lower()
                if choice == 'y':
                    return files["product"]["found"], files["kitchen"]["found"]
                else:
                    print("等待重新检测...")
                    input("按回车键继续...")
            else:
                print("\n" + "="*50)
                print("缺少以下必要文件：")
                for name in missing_files:
                    print(f"- {name}")
                print("\n文件命名要求：")
                print("1. 产品统计表：文件名需包含'产品统计表'")
                print("2. 厨房用表：文件名需包含'厨房'")
                print("="*50)
                input("请放置文件后按回车键重新扫描...")

    def read_product_data(self, filepath):
        """读取产品统计表数据"""
        wb = load_workbook(filepath, data_only=True)
        data = {}

        # 读取总表数据
        try:
            total_ws = wb['总表']
            ranges = [(3, 11), (13, 16), (30, 30), (32, 38)]
            for r in ranges:
                for row in range(r[0], r[1] + 1):
                    product = self.normalize_name(self.get_merged_value(total_ws, f'B{row}'))
                    value = self.safe_convert(self.get_merged_value(total_ws, f'E{row}'))
                    if product:
                        data[product] = value
        except KeyError:
            print("警告：未找到【总表】，跳过总表数据读取")

        # 读取用料表数据（修改后的范围）
        try:
            material_ws = wb['用料表']
            ranges = [(3, 9), (13, 13)]
            for start, end in ranges:
                for row in range(start, end + 1):
                    product = self.normalize_name(self.get_merged_value(material_ws, f'B{row}'))
                    value = self.safe_convert(self.get_merged_value(material_ws, f'E{row}'))
                    if product:
                        data[product] = value
        except KeyError:
            print("警告：未找到【用料表】，跳用料表数据读取")

        return data

    def read_kitchen_data(self, filepath):
        """读取厨房用表数据（跳过36行）"""
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        data = {}

        for row in range(5, 39):  # 对应Excel行号5-38
            if row == 37:  # 跳过37行
                continue
                
            product = self.normalize_name(self.get_merged_value(ws, f'B{row}'))
            value = self.safe_convert(self.get_merged_value(ws, f'F{row}'))
            if product:
                data[product] = value
        return data

    def get_merged_value(self, ws, cell_address):
        """获取合并单元格的值"""
        cell = ws[cell_address]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value

    def normalize_name(self, name):
        """统一产品名称格式"""
        if name is None:
            return ""
        name = str(name).strip()
        replacements = {
            '无糖酸奶': '零蔗糖酸奶',
            '蔓越莓酸奶': '蔓越莓胶原酸奶',
            '冷萃罐罐': '冷萃酸奶罐罐',
        #    '无糖试吃': '零蔗糖品尝',
            '半口奶酪品尝': '半口品尝',
            '奶皮子品尝': '开心果双皮奶品尝',
            '原味冷萃成品': '原味冷萃半成品'
        }
        return replacements.get(name, name)

    def safe_convert(self, value):
        """安全数值转换（处理公式和特殊格式）"""
        try:
            if value is None:
                return 0.0
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str) and value.startswith('='):
                return 0.0  # 需要提前计算保存文件
            cleaned = re.sub(r'[^\d\.]', '', str(value))
            return float(cleaned) if cleaned else 0.0
        except Exception as e:
            print(f"数值转换警告：{value} 无法转换，已视为0")
            return 0.0

    def compare_data(self):
        """主对比流程（包含特殊规则）"""
        try:
            product_file, kitchen_file = self.auto_detect_files()
            
            print("\n正在读取产品统计表数据...")
            self.product_data = self.read_product_data(product_file)
            
            print("正在读取厨房用表数据...")
            self.kitchen_data = self.read_kitchen_data(kitchen_file)
            
            print("\n开始数据对比...")
            differences = []

            # ================= 特殊对比规则 =================
            # 规则1：炒酸奶汇总对比
            fried_yogurt_total = 0.0
            fried_products = []
            for product, value in self.kitchen_data.items():
                if '炒酸奶' in product and '全家福' not in product:
                    fried_yogurt_total += value
                    fried_products.append(product)
            
            family_banquet = self.product_data.get('全家福炒酸奶（10块）', 0.0)
            if abs(fried_yogurt_total - family_banquet) > 0.01:
                differences.append({
                    'product': '[汇总]炒酸奶总量',
                    'type': '数值差异',
                    'product_value': family_banquet,
                    'kitchen_value': fried_yogurt_total,
                    'detail': f"厨房表({'+'.join(fried_products)})"
                })

            # 规则2：无糖系列特殊换算
            ws_no_sugar = self.kitchen_data.get('无糖品尝', 0.0)
            ws_tasting = self.kitchen_data.get('无糖试吃', 0.0) * 22  # 1:22比例换算
            product_target = self.product_data.get('零蔗糖品尝', 0.0)
            
            if abs((ws_no_sugar + ws_tasting) - product_target) > 0.01:
                differences.append({
                    'product': '[特殊换算]零蔗糖品尝',
                    'type': '数值差异',
                    'product_value': product_target,
                    'kitchen_value': ws_no_sugar + ws_tasting,
                    'detail': f"厨房表计算值({ws_no_sugar} + {self.kitchen_data.get('无糖试吃',0)}×22)"
                })

            # ================= 常规产品对比 =================
            excluded_products = {
                '全家福炒酸奶（10块）', '零蔗糖品尝',
                '无糖品尝', '无糖试吃'  # 排除已特殊处理的产品
            }
            
            # 产品统计表 → 厨房用表对比
            for product, prod_value in self.product_data.items():
                if product in excluded_products:
                    continue
                
                kitchen_value = self.kitchen_data.get(product)
                
                if kitchen_value is None:
                    differences.append({
                        'product': product,
                        'type': '产品缺失',
                        'product_value': prod_value,
                        'kitchen_value': None
                    })
                elif abs(prod_value - kitchen_value) > 0.01:
                    differences.append({
                        'product': product,
                        'type': '数值差异',
                        'product_value': prod_value,
                        'kitchen_value': kitchen_value
                    })

            # 厨房用表 → 产品统计表多余产品检查
            for product in self.kitchen_data:
                if product not in self.product_data and \
                   product not in excluded_products and \
                   '炒酸奶' not in product:
                    differences.append({
                        'product': product,
                        'type': '多余产品',
                        'product_value': None,
                        'kitchen_value': self.kitchen_data[product]
                    })

            # ================= 结果输出 =================
            if not differences:
                print("\n对比结果：所有数据一致！")
                return  # 直接返回，不执行后续代码            
            else:
                print(f"\n发现 {len(differences)} 处差异：")
                print("-"*70)
                for diff in differences:
                    if diff['product'].startswith('[汇总]'):
                        print(f"汇总对比：{diff['product'][3:]}")
                        print(f"产品统计表值：{diff['product_value']}")
                        print(f"厨房用表值：{diff['kitchen_value']} {diff['detail']}")
                        print(f"差值：{abs(diff['product_value'] - diff['kitchen_value']):.2f}")
                        print("-"*70)
                    elif diff['product'].startswith('[特殊换算]'):
                        print(f"特殊换算：{diff['product'][5:]}")
                        print(f"产品统计表值：{diff['product_value']}")
                        print(f"厨房用表计算值：{diff['kitchen_value']} ({diff['detail']})")
                        print(f"差值：{abs(diff['product_value'] - diff['kitchen_value']):.2f}")
                        print("-"*70)
                    else:
                        print(f"产品：{diff['product']}")
                        print(f"差异类型：{diff['type']}")
                        if diff['product_value'] is not None:
                            print(f"产品统计表值：{diff['product_value']}")
                        if diff['kitchen_value'] is not None:
                            print(f"厨房用表值：{diff['kitchen_value']}")
                        if diff['type'] == '数值差异':
                            print(f"差值：{abs(diff['product_value'] - diff['kitchen_value']):.2f}")
                        print("-"*70)

                # 生成报告
                self.generate_report(differences, product_file)

        except Exception as e:
            print(f"\n[处理失败] 发生错误：{str(e)}")

    def generate_report(self, differences, original_path):
        """生成差异报告（增强版）"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "差异报告"

        # 表头
        headers = ["对比类型", "差异说明", "产品统计表值", "厨房用表值", "差值"]
        ws.append(headers)

        # 数据行
        for diff in differences:
            # 处理汇总类差异
            if diff['product'].startswith('[汇总]'):
                row = [
                    "汇总对比",
                    f"{diff['product'][3:]} {diff['detail']}",
                    diff['product_value'],
                    diff['kitchen_value'],
                    abs(diff['product_value'] - diff['kitchen_value'])
                ]
            # 处理特殊换算
            elif diff['product'].startswith('[特殊换算]'):
                row = [
                    "特殊换算",
                    f"{diff['product'][5:]} {diff['detail']}",
                    diff['product_value'],
                    diff['kitchen_value'],
                    abs(diff['product_value'] - diff['kitchen_value'])
                ]
            # 处理常规差异
            else:
                diff_value = "/"
                if diff['type'] == '数值差异':
                    diff_value = abs(diff['product_value'] - diff['kitchen_value'])
                row = [
                    diff['type'],
                    diff['product'],
                    diff['product_value'] if diff['product_value'] is not None else "/",
                    diff['kitchen_value'] if diff['kitchen_value'] is not None else "/",
                    diff_value
                ]
            ws.append(row)

        # 文件名处理
        original_name = os.path.basename(original_path)
        new_name = original_name.replace(".xlsx", "_对比结果.xlsx")
        save_path = os.path.join(os.path.dirname(original_path), new_name)
        
        # 处理重名
        counter = 1
        while os.path.exists(save_path):
            new_name = original_name.replace(".xlsx", f"_对比结果({counter}).xlsx")
            save_path = os.path.join(os.path.dirname(original_path), new_name)
            counter += 1

        wb.save(save_path)
        print(f"\n差异报告已生成：{save_path}")

if __name__ == "__main__":
    print("="*50)
    print("Excel数据对比工具 V2.0")
    print("更新说明：")
    print("- 支持炒酸奶汇总对比")
    print("- 新增无糖产品特殊换算规则")
    print("- 优化数据读取范围和名称标准化")
    print("="*50)
    
    comparator = ExcelComparator()
    try:
        comparator.compare_data()
    except KeyboardInterrupt:
        print("\n操作已取消")
    except Exception as e:
        print(f"\n程序异常终止：{str(e)}")
    finally:
        os.system('pause' if os.name == 'nt' else 'read -p "按任意键退出..."')
