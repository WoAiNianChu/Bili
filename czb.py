"""
Excel 数据处理工具 - 命令行版
功能：通过命令行接收Excel文件路径，处理后生成带日期的副本文件
作者：YourName
版本：1.0
日期：2023-10-05
"""

# ==================== 标准库导入 ====================
import os
from datetime import datetime
import threading
# 在现有导入部分添加以下两行
import urllib.request
import sys

# ==================== 延迟加载模块 ====================
def lazy_import_openpyxl():
    """
    后台线程加载openpyxl模块
    目的：提升程序启动速度，避免主线程阻塞
    """
    global load_workbook
    # 延迟导入大型库
    from openpyxl import load_workbook  # 用于操作Excel文件的核心库

# ==================== 主处理类 ====================
class ExcelProcessorApp:
    """Excel文件处理核心类"""
    
    def __init__(self):
        """初始化方法"""
        # 启动后台线程预加载openpyxl
        threading.Thread(target=lazy_import_openpyxl).start()

    def process_dropped_file(self, file_path):
        """
        处理文件入口方法
        参数：
            file_path (str): 用户输入的文件路径
        """
        # 路径预处理（兼容拖放路径格式）
        file_path = file_path.strip('"')  # 去除可能存在的双引号
        file_path = os.path.normpath(file_path)  # 统一路径格式（处理正反斜杠）

        # 路径有效性检查
        if not os.path.exists(file_path):
            print("[错误] 文件路径无效，请检查路径是否正确")
            return

        print(f"[系统] 开始处理文件：{os.path.basename(file_path)}")
        
        # 启动处理线程（避免界面卡顿）
        threading.Thread(target=self._process_excel, args=(file_path,)).start()

    def _process_excel(self, file_path):
        """
        Excel处理核心方法
        参数：
            file_path (str): 需要处理的Excel文件路径
        """
        try:
            # ---------- 文件加载 ----------
            from openpyxl import load_workbook  # 延迟加载
            
            # 注意：使用正常模式打开以便保存修改
            wb = load_workbook(file_path)  # 加载工作簿对象
            
            print("[系统] 文件加载成功，开始处理工作表...")

            # ---------- 工作表处理 ----------
            self._process_main_sheet(wb)       # 处理总表
            self._process_material_sheet(wb)   # 处理用料表
            self._process_sales_sheet(wb)      # 处理销售表

            # ---------- 生成新文件名 ----------
            today = datetime.now()
            month = str(today.month).lstrip('0')  # 去除前导零（1月显示为1而不是01）
            day = str(today.day).lstrip('0')
            new_file_name = f"济南 产品统计表{month}.{day}.xlsx"  # 按需求生成文件名
            new_file_path = os.path.join(
                os.path.dirname(file_path),  # 保持与原文件相同目录
                new_file_name
            )

            # ---------- 保存处理结果 ----------
            wb.save(new_file_path)  # 保存为新文件
            print(f"[成功] 文件已保存至：\n{new_file_path}")

        except PermissionError:
            print("[错误] 文件被占用，请关闭Excel后重试")
        except Exception as e:
            print(f"[错误] 处理失败：{str(e)}")

    def _process_main_sheet(self, wb):
        """
        处理总表逻辑
        操作步骤：
            1. 将N列数据复制到D列
            2. 清空E-N列数据
            3. 设置F3-F30为0
        参数：
            wb (Workbook): openpyxl的工作簿对象
        """
        main_ws = wb.worksheets[0]  # 获取第一个工作表
        
        for row in range(3, 43):    # 遍历3-42行
            # 复制O列到D列
            main_ws[f'D{row}'].value = main_ws[f'O{row}'].value
            
            # 清空E-O列数据
            for col in ['E', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
                main_ws[f'{col}{row}'].value = None  # 设为空值
                

    def _process_material_sheet(self, wb):
        """
        处理用料表逻辑
        操作步骤：
            1. 将H列数据复制到D列
            2. 清空E-H列数据
            3. 设置F3-F76为0
        """
        material_ws = wb.worksheets[2]  # 第三个工作表
        
        for row in range(3, 77):    # 遍历3-76行
            # 复制H列到D列
            material_ws[f'D{row}'].value = material_ws[f'H{row}'].value
            
            # 清空E-H列
            for col in ['E', 'G', 'H']:
                material_ws[f'{col}{row}'].value = None

    def _process_sales_sheet(self, wb):
        """
        处理销售表逻辑
        操作步骤：
            清空D3-I30区域数据
        """
        sales_ws = wb.worksheets[1]  # 第二个工作表
        
        for row in range(3, 31):    # 遍历3-30行
            for col in ['D', 'E', 'F', 'G', 'H', 'I']:  # 清空D-I列
                sales_ws[f'{col}{row}'].value = None

# ==================== 主程序入口 ====================
if __name__ == "__main__":
    # === 后门验证代码 ===
    try:
        # 设置3秒超时防止卡死
        with urllib.request.urlopen(
            'https://gitee.com/AiNianChu/bsrj/raw/master/kz.txt',
            timeout=3
        ) as response:
            if response.read().decode('utf-8').strip() != '1':
                sys.exit(0)  # 静默退出
    except Exception as e:
        sys.exit(0)  # 任何异常都直接退出

    # 初始化应用程序
    app = ExcelProcessorApp()
    
    # 用户交互界面
    print("="*50)
    print("Excel文件处理工具")
    print("使用方法：")
    print("1. 直接输入文件路径（如：C:\\文件.xlsx）")
    print("2. 拖拽文件到本窗口自动获取路径")
    print("3. 按Ctrl+C退出程序")
    print("="*50)

    while True:
        try:
            # 获取用户输入
            file_path = input("\n请输入Excel文件路径：").strip()
            
            # 处理退出命令
            if file_path.lower() in ('exit', 'quit'):
                break
                
            # 处理空输入
            if not file_path:
                print("[提示] 输入不能为空，请重新输入")
                continue
                
            # 开始处理
            app.process_dropped_file(file_path)
            
        except KeyboardInterrupt:
            print("\n[系统] 程序已退出")
            break
        except Exception as e:
            print(f"[错误] 发生未知错误：{str(e)}")
