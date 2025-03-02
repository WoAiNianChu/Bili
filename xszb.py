import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class SalesDataUpdater:
    def __init__(self):
        self.stat_file = None
        self.total_file = None

    def auto_detect_files(self):
        """自动检测文件并显示详细信息"""
        required_files = {
            "product_stat": {"patterns": ["产品统计表"], "found": None},
            "sales_total": {"patterns": ["产品销售总表"], "found": None}
        }

        while True:
            os.system('cls' if os.name == 'nt' else 'clear')
            print("正在扫描目录...\n")
            current_dir = os.getcwd()
            
            # 扫描并获取最新文件
            for filename in os.listdir(current_dir):
                if not filename.endswith('.xlsx'):
                    continue
                filepath = os.path.join(current_dir, filename)
                
                # 产品统计表检测
                if all(p in filename for p in required_files["product_stat"]["patterns"]):
                    self.update_file_info(required_files["product_stat"], filepath)
                
                # 销售总表检测
                if all(p in filename for p in required_files["sales_total"]["patterns"]):
                    self.update_file_info(required_files["sales_total"], filepath)

            # 验证文件状态
            missing = [k for k, v in required_files.items() if not v["found"]]
            if not missing:
                self.show_file_info(required_files)
                if input("\n是否确认使用这些文件？(Y/N): ").lower() == 'y':
                    self.stat_file = required_files['product_stat']['found']
                    self.total_file = required_files['sales_total']['found']
                    return
            else:
                print("\n缺少以下文件：" + ", ".join(
                    ["产品统计表" if k == "product_stat" else "销售总表" for k in missing]
                ))
                input("请放置文件后按回车键重新扫描...")

    def update_file_info(self, file_info, new_path):
        """更新文件信息并保留最新版本"""
        if not file_info["found"] or \
           os.path.getmtime(new_path) > os.path.getmtime(file_info["found"]):
            file_info["found"] = new_path
            file_info["mtime"] = datetime.fromtimestamp(os.path.getmtime(new_path))

    def show_file_info(self, files):
        """显示文件详细信息"""
        print("   检测到最新文件：")
        stat = files["product_stat"]
        print(f"产品统计表：{os.path.basename(stat['found'])}")
        print(f"最后修改时间：{stat['mtime'].strftime('%Y-%m-%d %H:%M:%S')}")
        
        total = files["sales_total"]
        print(f"销售总表：{os.path.basename(total['found'])}")
        print(f"最后修改时间：{total['mtime'].strftime('%Y-%m-%d %H:%M:%S')}")

    def get_target_column(self):
        """根据当前日期计算销售数据列（销售数据列为：起始E列，每日占2列）"""
        today = datetime.now()
        current_day = today.day
        data_col_num = 5 + 2 * (current_day - 1)  # E列开始（E对应5）
        # 限制最大列为66（即BO列）
        return get_column_letter(min(data_col_num, 66))

    def get_remark_column(self):
        """根据当前日期计算备注列（销售数据列+1）"""
        today = datetime.now()
        current_day = today.day
        
        # 计算销售数据列号（每日占2列，从E列开始）
        data_col_num = 5 + 2 * (current_day - 1)  # E列开始
        # 备注列为销售数据列+1
        remark_col_num = data_col_num + 1
        
        # 保护逻辑：最大列数不超过BO（66列）
        return get_column_letter(min(remark_col_num, 66))

    def copy_remarks(self):
        """执行备注数据复制操作"""
        stat_wb = load_workbook(self.stat_file, data_only=True)
        total_wb = load_workbook(self.total_file)
        
        try:
            stat_ws = stat_wb["总表"]
            total_ws = total_wb["总表"]
            
            target_col = self.get_remark_column()
            today_str = datetime.now().strftime("%Y/%m/%d")
            
            print(f"\n   正在更新备注数据到 [{today_str}] 列...")
            
            # Q3-Q42 -> 总表4-43行
            for src_row, dst_row in zip(range(3, 43), range(4, 44)):
                cell_value = stat_ws[f'Q{src_row}'].value
                # 数据清洗
                if cell_value is None:
                    cell_value = ""
                elif isinstance(cell_value, float) and cell_value.is_integer():
                    cell_value = int(cell_value)
                total_ws[f'{target_col}{dst_row}'] = cell_value
            
            total_wb.save(self.total_file)
            print(f"   备注数据更新成功！目标列：{target_col}")
            
        finally:
            stat_wb.close()
            total_wb.close()

    def copy_data(self):
        """执行数据复制操作"""
        stat_wb = load_workbook(self.stat_file, data_only=True)
        total_wb = load_workbook(self.total_file)
        
        try:
            stat_ws = stat_wb["总表"]
            total_ws = total_wb["总表"]
            
            # 销售数据列
            target_col = self.get_target_column()
            today_str = datetime.now().strftime("%Y/%m/%d")
            
            print(f"\n   正在更新销售数据到 [{today_str}] 列...")
            
            # 第一部分：F3-F30 -> 总表4-31
            for src_row, dst_row in zip(range(3, 31), range(4, 32)):
                cell_value = stat_ws[f'F{src_row}'].value
                total_ws[f'{target_col}{dst_row}'] = cell_value
                
            # 第二部分：J31 -> 总表32
            for src_row, dst_row in zip(range(31, 32), range(32, 33)):
                cell_value = stat_ws[f'J{src_row}'].value
                total_ws[f'{target_col}{dst_row}'] = cell_value

            # 第三部分：G32-G42 -> 总表33-43
            for src_row, dst_row in zip(range(32, 43), range(33, 44)):
                cell_value = stat_ws[f'G{src_row}'].value
                total_ws[f'{target_col}{dst_row}'] = cell_value
            
            total_wb.save(self.total_file)
            print(f"   销售数据更新成功！目标列：{target_col}")
            
        finally:
            stat_wb.close()
            total_wb.close()

    def run(self):
        self.auto_detect_files()
        self.copy_data()    # 原有销售数据复制
        self.copy_remarks()  # 新增备注数据复制

if __name__ == "__main__":
    print("=" * 50)
    print("📥 销售数据同步工具 V2.1")
    print("功能特性：")
    print("- 自动识别最新版本文件")
    print("- 销售数据与备注同步更新")
    print("- 双列定位系统（销售列+备注列）")
    print("- 直接覆盖原文件保存")
    print("=" * 50)
    
    try:
        SalesDataUpdater().run()
    except Exception as e:
        print(f"\n❌ 操作失败：{str(e)}")
    finally:
        os.system('pause' if os.name == 'nt' else 'read -p "按任意键退出..."')
