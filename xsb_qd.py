import os
import re
from datetime import datetime
import threading
from openpyxl import load_workbook

def set_cell_value(ws, cell_address, value):
    """安全设置单元格值（处理合并单元格）"""
    target_cell = ws[cell_address]

    # 检查合并单元格
    for merged_range in ws.merged_cells.ranges:
        if target_cell.coordinate in merged_range:
            top_left_cell = ws.cell(
                row=merged_range.min_row, 
                column=merged_range.min_col
            )
            if target_cell.coordinate == top_left_cell.coordinate:
                top_left_cell.value = value
            return

    # 普通单元格直接赋值
    target_cell.value = value

class ExcelProcessorApp:
    """Excel文件处理核心类"""

    def __init__(self):
        """初始化时预加载依赖"""
        threading.Thread(target=self.lazy_import_openpyxl).start()

    def lazy_import_openpyxl(self):
        """延迟加载openpyxl模块"""
        global load_workbook
        from openpyxl import load_workbook

    def auto_detect_files(self):
        """自动检测同目录下的三个必要文件"""
        while True:
            os.system('cls' if os.name == 'nt' else 'clear')
            print("正在扫描目录...")
            current_dir = os.getcwd()
            files = {
                "groupon": {"pattern": r'^\d{4}-\d{2}-\d{2}.*', "found": None, "name": "团购表"},
                "product": {"pattern": "产品统计表", "found": None, "name": "产品统计表"},
                "ranking": {"pattern": "商品排行报表", "found": None, "name": "商品排行报表"}
            }

            # 扫描目录并匹配文件
            for filename in os.listdir(current_dir):
                filepath = os.path.join(current_dir, filename)
                if not os.path.isfile(filepath):
                    continue

                # 团购表匹配
                if re.match(files["groupon"]["pattern"], filename):
                    if not files["groupon"]["found"] or \
                       os.path.getmtime(filepath) > os.path.getmtime(files["groupon"]["found"]):
                        files["groupon"]["found"] = filepath

                # 产品统计表匹配
                if files["product"]["pattern"] in filename and "_已处理" not in filename:
                    if not files["product"]["found"] or \
                       os.path.getmtime(filepath) > os.path.getmtime(files["product"]["found"]):
                        files["product"]["found"] = filepath

                # 商品排行报表匹配
                if files["ranking"]["pattern"] in filename:
                    if not files["ranking"]["found"] or \
                       os.path.getmtime(filepath) > os.path.getmtime(files["ranking"]["found"]):
                        files["ranking"]["found"] = filepath

            # 检查文件是否齐全
            missing_files = [f["name"] for f in files.values() if not f["found"]]
            if not missing_files:
                # 显示检测结果
                print("\n" + "="*50)
                print("检测到以下最新文件：")
                for f_type in files.values():
                    mtime = datetime.fromtimestamp(os.path.getmtime(f_type["found"]))
                    print(f"[{f_type['name']}]")
                    print(f"文件名：{os.path.basename(f_type['found'])}")
                    print(f"修改时间：{mtime.strftime('%Y-%m-%d %H:%M:%S')}\n")
                print("="*50)

                # 用户确认
                choice = input("是否开始处理？(Y/N): ").strip().lower()
                if choice == 'y':
                    return (
                        files["ranking"]["found"],  # 商品排行报表
                        files["product"]["found"],  # 产品统计表
                        files["groupon"]["found"]   # 团购表
                    )
                else:
                    print("等待重新检测...")
                    input("按回车键继续...")
            else:
                print("\n" + "="*50)
                print("缺少以下必要文件：")
                for name in missing_files:
                    print(f"- {name}")
                print("\n文件命名要求：")
                print("1. 团购表：以日期开头（如：2024-03-15_团购数据.xlsx）")
                print("2. 产品统计表：文件名需包含'产品统计表'且不含'_已处理'")
                print("3. 商品排行报表：文件名需包含'商品排行报表'")
                print("="*50)
                input("请放置文件后按回车键重新扫描...")

    def process_files(self):
        """主处理流程"""
        try:
            # 自动获取文件路径
            ranking_file, product_file, groupon_file = self.auto_detect_files()

            # 处理商品排行报表
            print(f"\n正在处理商品排行报表：{os.path.basename(ranking_file)}")
            ranking_wb = load_workbook(ranking_file)
            product_sales, e_sales = self.merge_product_sales(ranking_wb.active)

            # 处理团购报表
            print(f"正在处理美团团购报表：{os.path.basename(groupon_file)}")
            groupon_wb = load_workbook(groupon_file)
            groupon_sales = self.process_groupon_sales(groupon_wb.active)

            # 处理产品统计表
            print(f"正在处理产品统计表：{os.path.basename(product_file)}")
            product_wb = load_workbook(product_file)
            product_ws = product_wb["销售表"]

            # 更新数据
            self.update_product_sales(product_ws, product_sales, e_sales, groupon_sales)

            # 生成基于当前日期的新文件名
            today = datetime.now()
            base_name = f"济南 产品统计表{today.month}-{today.day}"
            new_filename = f"{base_name}.xlsx"
            new_file_path = os.path.join(os.path.dirname(product_file), new_filename)

            # 处理文件重名
            if os.path.exists(new_file_path):
                base_name += "_已处理"
                new_filename = f"{base_name}.xlsx"
                new_file_path = os.path.join(os.path.dirname(product_file), new_filename)

                # 处理多次重复
                counter = 1
                while os.path.exists(new_file_path):
                    new_filename = f"{base_name}({counter}).xlsx"
                    new_file_path = os.path.join(os.path.dirname(product_file), new_filename)
                    counter += 1

            product_wb.save(new_file_path)
            print(f"\n[成功] 文件已保存至：{new_file_path}")

        except Exception as e:
            print(f"\n[处理失败] 发生错误：{str(e)}")
        finally:
            input("\n处理完成，按回车键退出...")



    def merge_product_sales(self, ws):
        """处理商品排行报表数据"""
        product_sales = {}
        e_sales = {'饿了么外卖': {}, '美团外卖': {}}

        for row in range(2, ws.max_row + 1):
            # 原始数据提取
            raw_name = ws[f'C{row}'].value or ""
            original_quantity = self.parse_quantity(ws[f'F{row}'].value)
            e_type = ws[f'E{row}'].value or ""
            quantity = original_quantity
            product_name = None

            # 处理鲜牛奶的特殊情况（新增核心逻辑）
            if "鲜" in raw_name and "牛奶" in raw_name:
                # 检查包/次数量
                match = re.search(r'(\d+)(包|次|份)', raw_name)
                if match:
                    # 提取数量并计算实际销量
                    multiplier = int(match.group(1))
                    quantity = original_quantity * multiplier
                # 无论是否匹配到包/次都归类为鲜牛奶
                product_name = "鲜牛奶"
            else:
                # 处理其他商品逻辑
                if "炒酸奶" in raw_name:
                    quantity = original_quantity * 10  # 炒酸奶按10块/份计算
                # 标准化商品名称
                product_name = self.normalize_product_name(raw_name)

            # 销量累加
            product_sales[product_name] = product_sales.get(product_name, 0) + quantity

            # 渠道分类（关键修改点）
            if "未映射饿了么" in str(e_type):
                e_sales['饿了么外卖'][product_name] = e_sales['饿了么外卖'].get(product_name, 0) + quantity
            elif "未映射美团" in str(e_type):
                e_sales['美团外卖'][product_name] = e_sales['美团外卖'].get(product_name, 0) + quantity

        return product_sales, e_sales

    def process_groupon_sales(self, ws):
        """处理团购报表数据（已更新鲜牛奶逻辑）"""
        headers = {cell.value: cell.column_letter for cell in ws[1]}
        required_cols = ['核销时间', '商品名称', '验证门店']
        if missing := [col for col in required_cols if col not in headers]:
            print(f"[错误] 缺少必要列：{', '.join(missing)}")
            return {}

        groupon_sales = {}
        today = datetime.now().date()

        for row in range(2, ws.max_row + 1):
            # 日期过滤
            date_val = ws[f"{headers['核销时间']}{row}"].value
            if not isinstance(date_val, datetime):
                try:
                    date_val = datetime.strptime(date_val, "%Y-%m-%d %H:%M:%S")
                except (ValueError, TypeError):
                    continue
            if date_val.date() != today:
                continue

            # 门店过滤
            store = ws[f"{headers['验证门店']}{row}"].value or ""
            if "济南" not in store:
                continue

            # 商品处理（新增核心逻辑）
            raw_name = ws[f"{headers['商品名称']}{row}"].value or ""
            original_quantity = 1  # 团购每条记录默认1次核销
            quantity = original_quantity
            product_name = None

        # 处理鲜牛奶
            if "鲜" in raw_name and "牛奶" in raw_name:
                match = re.search(r'(\d+)(包|次|份)', raw_name)
                if match:
                    multiplier = int(match.group(1))
                    quantity = original_quantity * multiplier
                product_name = "鲜牛奶"
            else:
                # 处理炒酸奶
                if "炒酸奶" in raw_name:
                    quantity = original_quantity * 10
                # 标准化商品名称
                product_name = self.normalize_product_name(raw_name)

            # 累加销量
            groupon_sales[product_name] = groupon_sales.get(product_name, 0) + quantity

        return groupon_sales

    def update_product_sales(self, ws, product_sales, e_sales, groupon_sales):
        """更新产品统计表"""
        for row in range(2, ws.max_row + 1):
            product_name = ws[f'B{row}'].value

            # 安全写入函数
            def safe_write(col, value):
                if value is not None and value != 0:  # 只写入非零值
                    set_cell_value(ws, f"{col}{row}", value)

            # 获取销量数据
            total_sales = product_sales.get(product_name, 0)
            eleme_sales = e_sales['饿了么外卖'].get(product_name, 0)
            meituan_sales = e_sales['美团外卖'].get(product_name, 0)
            groupon_sales_value = groupon_sales.get(product_name, 0)
            # 直接计算零售（D列 - 各渠道销量）
            retail_sales = total_sales - (groupon_sales_value + eleme_sales + meituan_sales)

            # 更新各列数据
            safe_write('D', total_sales)  # 小条
            safe_write('H', eleme_sales)  # 饿了么
            safe_write('G', meituan_sales)  # 美团
            safe_write('F', groupon_sales_value)  # 团购
            safe_write('E', retail_sales)  # 直接填零售数据

            # 添加小条计算公式（D列 = 各渠道之和）
            if 3 <= row <= 30:  # 根据实际数据范围调整
                formula = f"=SUM(E{row}:I{row})"
                set_cell_value(ws, f'D{row}', formula)

    def normalize_product_name(self, name):
        """商品名称标准化（增强版）"""
        # 清理特殊符号
        cleaned = re.sub(r'【.*?】|\(.*?\)|\d+块|\d+份', '', str(name)).strip()

        # 定义标准化规则（优先级从高到低）
        rules = [
            # 炒酸奶相关
            (lambda x: "全家福" in x and "炒酸奶" in x, "全家福炒酸奶"),
            (lambda x: "炒酸奶" in x and "10块" in x, "全家福炒酸奶"),
            (lambda x: "炒酸奶" in x, "全家福炒酸奶"),

            # 鲜牛乳系列
            (lambda x: "草莓" in x and "鲜牛乳" in x, "草莓冷萃鲜牛乳"),
            (lambda x: "开心果" in x and "鲜牛乳" in x, "开心果冷萃鲜牛乳"),
            (lambda x: "抹茶" in x and "鲜牛乳" in x, "抹茶冷萃鲜牛乳"),
            (lambda x: ("芋泥" in x or "香芋" in x) and "鲜牛乳" in x, "香芋冷萃鲜牛乳"),

            # 冰淇淋系列
            (lambda x: ("鲜奶" in x or "牛奶" in x) and "冰淇淋" in x, "鲜奶冰淇淋"),
            (lambda x: ("酸奶" in x or "酸" in x) and "冰淇淋" in x, "酸奶冰淇淋"),       

            # 酸奶碗系列
            (lambda x: ("圣诞" in x or "草莓" in x) and "酸奶碗" in x, "酸奶碗—草莓"),
            (lambda x: ("希腊冷萃" in x or "开心果" in x) and "酸奶碗" in x, "酸奶碗—开心果能量"),    

            # 鸳鸯酸奶系列
            (lambda x: "草莓" in x and "鸳鸯" in x, "草莓鸳鸯酸奶"),
            (lambda x: "开心果" in x and "鸳鸯" in x, "开心果鸳鸯酸奶"),


            (lambda x: "蔓越莓" in x, "蔓越莓胶原酸奶"),
            (lambda x: "双蛋白" in x, "双蛋白酸奶"),
            (lambda x: "零蔗糖" in x, "零蔗糖酸奶"),
            (lambda x: "芝士" in x, "芝士酸奶"),
            (lambda x: "紫米" in x, "紫米酸奶"),
            (lambda x: "液体酸奶" in x, "液体酸奶"),
            (lambda x: "奶皮子" in x, "奶皮子奶酪"),

            # 其他商品
            (lambda x: "布丁" in x, "布丁"),
            (lambda x: "生巧" in x, "生巧可可牛奶"),
            (lambda x: "香蕉" in x, "香蕉牛奶"),
            (lambda x: "半口" in x, "半口奶酪"),
            (lambda x: "罐罐" in x, "冷萃酸奶罐罐"),
            (lambda x: "双皮奶" in x and "原味" not in x, "果味双皮奶"),
        ]

        # 应用匹配规则
        for condition, standardized in rules:
            if condition(cleaned):
                return standardized
        return cleaned  # 默认返回清理后的名称

    def parse_quantity(self, value):
        """通用数量解析"""
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0

if __name__ == "__main__":
    app = ExcelProcessorApp()
    print("="*50)
    print("Excel自动化处理工具 V2.3")
    print("功能特点：")
    print("- 自动识别最新文件（支持拖放文件到目录）")
    print("- 智能处理鲜牛奶/炒酸奶等特殊商品")
    print("- 自动计算各渠道销售数据")
    print("="*50)
    print("操作提示：")
    print("1. 请确保三个文件在同一个目录")
    print("2. 程序会自动选择最新版本的文件")
    print("3. 按提示确认即可开始处理")
    print("="*50)

    try:
        app.process_files()
    except KeyboardInterrupt:
        print("\n操作已取消")
    except Exception as e:
        print(f"\n程序异常终止：{str(e)}")
    finally:
        os.system('pause' if os.name == 'nt' else 'read -p "按任意键退出..."')