import os
import re
from datetime import datetime
import threading
from openpyxl import load_workbook

def set_cell_value(ws, cell_address, value):
    """
    安全设置单元格值（自动处理合并单元格）
    :param ws: 工作表对象
    :param cell_address: 目标单元格地址（如"A1"）
    :param value: 要设置的值
    """
    target_cell = ws[cell_address]
    
    # 检查是否在合并区域内
    for merged_range in ws.merged_cells.ranges:
        if target_cell.coordinate in merged_range:
            # 找到合并区域的左上角单元格
            top_left_cell = ws.cell(
                row=merged_range.min_row, 
                column=merged_range.min_col
            )
            # 只在左上角单元格写入值
            if target_cell.coordinate == top_left_cell.coordinate:
                top_left_cell.value = value
            return
    
    # 非合并单元格直接写入
    target_cell.value = value

class ExcelProcessorApp:
    """Excel文件处理核心类"""

    def __init__(self):
        """初始化方法"""
        threading.Thread(target=self.lazy_import_openpyxl).start()

    def lazy_import_openpyxl(self):
        global load_workbook
        from openpyxl import load_workbook

    def process_files(self):
        # 文件路径获取和验证
        ranking_file_path = self.sanitize_path(input("请输入商品排行报表文件路径："))
        product_file_path = self.sanitize_path(input("请输入产品统计表文件路径："))
        groupon_file_path = self.sanitize_path(input("请输入美团团购报表文件路径："))

        if not all(os.path.exists(p) for p in [ranking_file_path, product_file_path, groupon_file_path]):
            print("[错误] 文件路径无效，请检查路径是否正确")
            input("按回车键退出...")
            return

        # 处理商品排行报表
        print(f"正在处理商品排行报表：{ranking_file_path}")
        ranking_wb = load_workbook(ranking_file_path)
        product_sales, e_sales = self.merge_product_sales(ranking_wb.active)

        # 处理团购报表
        print(f"正在处理美团团购报表：{groupon_file_path}")
        groupon_wb = load_workbook(groupon_file_path)
        groupon_sales = self.process_groupon_sales(groupon_wb.active)

        # 处理产品统计表
        print(f"正在处理产品统计表：{product_file_path}")
        product_wb = load_workbook(product_file_path)
        product_ws = product_wb["销售表"]

        # 更新数据并添加公式
        self.update_product_sales(product_ws, product_sales, e_sales, groupon_sales)

        # 保存文件
        today = datetime.now()
        new_file_path = os.path.join(
            os.path.dirname(product_file_path),
            f"济南 产品统计表{today.month}-{today.day}.xlsx"
        )
        product_wb.save(new_file_path)
        print(f"[成功] 文件已保存至：{new_file_path}")
        input("处理完成，按回车键退出...")

    def merge_product_sales(self, ws):
        product_sales = {}
        e_sales = {'饿了么外卖': {}, '美团外卖': {}}

        for row in range(2, ws.max_row + 1):
            # 原始数据处理
            raw_name = ws[f'C{row}'].value or ""
            quantity = self.parse_quantity(ws[f'F{row}'].value)
            e_type = ws[f'E{row}'].value or ""

            # 处理包数逻辑
            if "鲜牛奶" in raw_name and "包" in raw_name:
                if match := re.search(r'(\d+)包', raw_name):
                    quantity *= int(match.group(1))

            # 标准化名称
            product_name = self.normalize_product_name(raw_name)

            # 累计销量
            product_sales[product_name] = product_sales.get(product_name, 0) + quantity

            # 渠道分类
            if "未映射饿了么" in str(e_type):
                e_sales['饿了么外卖'][product_name] = e_sales['饿了么外卖'].get(product_name, 0) + quantity
            elif "未映射美团" in str(e_type):
                e_sales['美团外卖'][product_name] = e_sales['美团外卖'].get(product_name, 0) + quantity

        return product_sales, e_sales

    def process_groupon_sales(self, ws):
        headers = {cell.value: cell.column_letter for cell in ws[1]}
        required_cols = ['核销时间', '商品名称', '验证门店']
        if missing := [col for col in required_cols if col not in headers]:
            print(f"[错误] 缺少必要列：{', '.join(missing)}")
            return {}

        groupon_sales = {}
        today = datetime.now().date()

        for row in range(2, ws.max_row + 1):
            # 日期过滤
            date_val = ws[f"{headers['核销时间']}{row}"].value
            if not isinstance(date_val, datetime):
                try:
                    date_val = datetime.strptime(date_val, "%Y-%m-%d %H:%M:%S")
                except (ValueError, TypeError):
                    continue
            if date_val.date() != today:
                continue

            # 门店过滤
            store = ws[f"{headers['验证门店']}{row}"].value or ""
            if "济南" not in store:
                continue

            # 处理商品
            raw_name = ws[f"{headers['商品名称']}{row}"].value or ""
            quantity = 1
            if "鲜牛奶" in raw_name and "包" in raw_name:
                if match := re.search(r'(\d+)包', raw_name):
                    quantity = int(match.group(1))

            product_name = self.normalize_product_name(raw_name)
            groupon_sales[product_name] = groupon_sales.get(product_name, 0) + quantity

        return groupon_sales

    def update_product_sales(self, ws, product_sales, e_sales, groupon_sales):
        for row in range(2, ws.max_row + 1):
            product_name = ws[f'B{row}'].value

            # 安全写入函数
            def write_cell(col, value):
                if value is not None:
                    set_cell_value(ws, f"{col}{row}", value)

            # 更新数据
            write_cell('D', product_sales.get(product_name))
            write_cell('H', e_sales['饿了么外卖'].get(product_name))
            write_cell('G', e_sales['美团外卖'].get(product_name))
            write_cell('F', groupon_sales.get(product_name))

            # 添加公式（处理合并单元格）
            if 3 <= row <= 29 and ws[f'D{row}'].value is not None:
                formula = f"=D{row}-SUM(F{row}:I{row})"
                set_cell_value(ws, f'E{row}', formula)

    def normalize_product_name(self, name):
        """标准化商品名称（包含处理【】符号）"""
        # 去除【】及其中间内容
        cleaned = re.sub(r'【.*?】', '', str(name)).strip()
        
        rules = [
            (lambda x: "草莓" in x and "鲜牛乳" in x, "草莓冷萃鲜牛乳"),
            (lambda x: "开心果" in x and "鲜牛乳" in x, "开心果冷萃鲜牛乳"),
            (lambda x: "抹茶" in x and "鲜牛乳" in x, "抹茶冷萃鲜牛乳"),
            (lambda x: ("芋泥" in x or "香芋" in x) and "鲜牛乳" in x, "香芋冷萃鲜牛乳"),
            (lambda x: ("鲜奶" in x or "牛奶" in x) and "冰淇淋" in x, "鲜奶冰淇淋"),
            (lambda x: "蔓越莓" in x, "蔓越莓胶原酸奶"),
            (lambda x: "双蛋白" in x, "双蛋白酸奶"),
            (lambda x: "零蔗糖" in x, "零蔗糖酸奶"),
            (lambda x: "芝士" in x, "芝士酸奶"),
            (lambda x: "紫米" in x, "紫米酸奶"),
            (lambda x: "液体酸奶" in x, "液体酸奶"),
            (lambda x: "奶皮子" in x, "奶皮子酸奶酪"),
            (lambda x: "布丁" in x, "布丁"),
            (lambda x: "生巧" in x, "生巧可可牛奶"),
            (lambda x: "香蕉" in x, "香蕉牛奶"),
            (lambda x: "半口" in x, "半口奶酪"),
            (lambda x: "罐罐" in x, "冷萃酸奶罐罐"),
            (lambda x: "酸奶碗" in x, "酸奶碗—开心果能量"),
            (lambda x: "鲜牛奶" in x and "包" not in x, "鲜牛奶"),
            (lambda x: "双皮奶" in x and "原味" not in x, "果味双皮奶"),
        ]

        for condition, standardized in rules:
            if condition(cleaned):
                return standardized
        return cleaned

    def parse_quantity(self, value):
        """通用数量解析"""
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0

    def sanitize_path(self, raw_input):
        """路径清理"""
        quoted = re.findall(r'["\'](.*?)["\']', raw_input)
        path = quoted[-1].strip() if quoted else raw_input.strip(" &'\"")
        return path if os.path.exists(path) else raw_input.strip("'\"")

if __name__ == "__main__":
    app = ExcelProcessorApp()
    print("="*50)
    print("Excel文件处理工具")
    print("使用方法：")
    print("1. 请依次输入三个文件路径")
    print("2. 商品排行报表、产品统计表、美团团购报表")
    print("3. 按Ctrl+C可随时退出程序")
    print("="*50)

    try:
        app.process_files()
    except KeyboardInterrupt:
        print("\n[系统] 程序已退出")
    except Exception as e:
        print(f"[严重错误] 程序运行异常：{str(e)}")