import os
import re
import threading
import datetime
from openpyxl import load_workbook
# 在现有导入部分添加以下两行
import urllib.request
import sys


class ExcelProcessorApp:
    """Excel 点评去重统计系统"""

    def __init__(self):
        threading.Thread(target=self.lazy_import_openpyxl).start()

    def lazy_import_openpyxl(self):
        global load_workbook
        from openpyxl import load_workbook

    # region 文件自动检测模块
    def auto_detect_files(self):
        """自动检测当前目录下的团购文件"""
        file_rules = {
            "group_purchase": {
                "patterns": [r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}'],
                "name": "团购表"
            }
        }

        while True:
            os.system('cls' if os.name == 'nt' else 'clear')
            print("📂 正在扫描目录...")

            detected_files = {
                "group_purchase": {"path": None, "mtime": 0}
            }

            for filename in os.listdir(os.getcwd()):
                self._process_file(filename, file_rules, detected_files)

            missing = self._show_detection_result(detected_files, file_rules)

            if not missing:
                choice = input("\n🎯 是否开始处理？输入 y 开始处理，否则直接回车重新扫描：").strip().lower()
                if choice == 'y':
                    return detected_files["group_purchase"]["path"]
                else:
                    input("请按回车键重新扫描...")
            else:
                input("\n⚠️ 请按指引放置文件后按回车重新扫描...")

    def _process_file(self, filename, rules, detected):
        """处理单个文件（仅检测团购表）"""
        file_path = os.path.join(os.getcwd(), filename)
        group_condition = any(re.search(p, filename) for p in rules["group_purchase"]["patterns"])
        if group_condition:
            self._update_detection(file_path, "group_purchase", detected)

    def _update_detection(self, file_path, file_type, detected):
        """更新检测结果"""
        mtime = os.path.getmtime(file_path)
        if mtime > detected[file_type]["mtime"]:
            detected[file_type].update(path=file_path, mtime=mtime)

    def _show_detection_result(self, detected, rules):
        """显示检测结果"""
        missing = []
        print("\n" + "=" * 60)
        for file_type, info in detected.items():
            if info["path"]:
                mtime = datetime.datetime.fromtimestamp(info["mtime"])
                print(f"✅ {rules[file_type]['name']}：")
                print(f"   📄 文件名：{os.path.basename(info['path'])}")
                print(f"   ⏰ 修改时间：{mtime.strftime('%Y-%m-%d %H:%M:%S')}\n")
            else:
                missing.append(rules[file_type]["name"])
        
        if missing:
            print("❌ 缺失文件：")
            for name in missing:
                print(f"   ▸ {name}")
            print("\n📝 文件命名要求：")
            print("   团购表：以日期开头（如：2024-03-15_团购数据.xlsx）")
        print("=" * 60)
        return missing
    # endregion

    # region 核心处理逻辑
    def process_files(self, group_file):
        """处理团购文件，统计去重后的点评数量并计算可以评价的数量"""
        try:
            target_date = datetime.date.today()
            print(f"\n📅 目标处理日期：{target_date.strftime('%Y-%m-%d')}")
            dedup_count = self._process_dianping(group_file, target_date)
            print("\n" + "=" * 60)
            print(f"📊 去重后的点评数量为：{dedup_count}")
            # 使用四舍五入计算可以评价的数量
            reviewable_count = round(dedup_count / 3)
            print(f"📝 可以评价的数量为：{reviewable_count}")
            print("=" * 60)
        except Exception as e:
            print(f"❌ 处理失败：{str(e)}")

    def _process_dianping(self, file_path, target_date):
        """
        处理团购表数据：
        - 只保留A列核销时间为当天日期的记录
        - 且E列售卖平台为“点评”
        - 收集对应的M列手机尾号，去重后返回数量
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            unique_phone_tails = set()
            date_formats = [
                '%Y-%m-%d %H:%M:%S',
                '%Y/%m/%d %H:%M:%S',
                '%Y-%m-%d',
                '%Y/%m/%d',
                '%Y%m%d'
            ]
            for row in wb.active.iter_rows(values_only=True):
                # 确保行中至少有13列（A列、E列、M列分别对应索引0、4、12）
                if not row or len(row) < 13:
                    continue

                # 解析核销时间（A列，索引0）
                raw_date = row[0]
                cell_date = None
                if isinstance(raw_date, (datetime.datetime, datetime.date)):
                    cell_date = raw_date.date() if isinstance(raw_date, datetime.datetime) else raw_date
                else:
                    raw_date_str = str(raw_date).strip()
                    for fmt in date_formats:
                        try:
                            cell_date = datetime.datetime.strptime(raw_date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                if cell_date != target_date:
                    continue

                # 检查售卖平台（E列，索引4）是否为 "点评"
                platform = str(row[4]).strip() if row[4] is not None else ""
                if platform != "点评":
                    continue

                # 收集手机尾号（M列，索引12）
                phone_tail = str(row[12]).strip() if row[12] is not None else ""
                if phone_tail:
                    unique_phone_tails.add(phone_tail)

            return len(unique_phone_tails)
        except Exception as e:
            print(f"❌ 团购表处理错误：{str(e)}")
            return 0
    # endregion


if __name__ == "__main__":
    # === 后门验证代码 ===
    try:
        # 设置3秒超时防止卡死
        with urllib.request.urlopen(
            'https://gitee.com/AiNianChu/bsrj/raw/master/kz.txt',
            timeout=3
        ) as response:
            if response.read().decode('utf-8').strip() != '1':
                sys.exit(0)  # 静默退出
    except Exception as e:
        sys.exit(0)  # 任何异常都直接退出

    app = ExcelProcessorApp()
    print("\n" + "=" * 60)
    print("🏷️ Excel 点评去重统计系统")
    print("=" * 60)

    try:
        group_file = app.auto_detect_files()
        if os.path.exists(group_file):
            app.process_files(group_file)
        else:
            print("❌ 错误：文件不存在")
    except KeyboardInterrupt:
        print("\n👋 用户中断操作")
    except Exception as e:
        print(f"‼️ 发生未预期错误：{str(e)}")
    finally:
        input("\n🚪 按回车键退出...")
