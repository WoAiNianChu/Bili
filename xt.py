import os
import re
import threading
import datetime
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class ExcelProcessorApp:
    """Excel 文件处理最终版"""

    def __init__(self):
        threading.Thread(target=self.lazy_import_openpyxl).start()

    def lazy_import_openpyxl(self):
        global load_workbook
        from openpyxl import load_workbook

    # region ################### 路径处理模块 ###################
    def sanitize_path(self, raw_input):
        """智能清理用户输入的路径"""
        quoted_paths = re.findall(r'["\'](.*?)["\']', raw_input)
        return quoted_paths[-1].strip() if quoted_paths else raw_input.strip(" &'\"")

    # endregion

    # region ################### 文件自动检测模块 ###################
    def auto_detect_files(self):
        """自动检测同目录下的必要文件"""
        file_rules = {
            "payment_stats": {
                "patterns": ["支付方式收款统计", r'^收款统计_'],
                "exclude": ["_已处理"],
                "name": "支付统计表"
            },
            "group_purchase": {
                "patterns": [r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}'],
                "name": "团购表"
            }
        }

        while True:
            os.system('cls' if os.name == 'nt' else 'clear')
            print("📂 正在扫描目录...")
            
            detected_files = {
                "payment_stats": {"path": None, "mtime": 0},
                "group_purchase": {"path": None, "mtime": 0}
            }

            for filename in os.listdir(os.getcwd()):
                self._process_file(filename, file_rules, detected_files)

            missing = self._show_detection_result(detected_files, file_rules)
            
            if not missing:
                choice = input("\n🎯 是否开始处理？(Y/N): ").lower()
                if choice == 'y':
                    return (
                        detected_files["payment_stats"]["path"],
                        detected_files["group_purchase"]["path"]
                    )
            else:
                input("\n⚠️ 请按指引放置文件后按回车重新扫描...")

    def _process_file(self, filename, rules, detected):
        """处理单个文件"""
        file_path = os.path.join(os.getcwd(), filename)
        
        # 支付统计表检测
        payment_condition = (
            any(p in filename for p in rules["payment_stats"]["patterns"]) 
            and not any(e in filename for e in rules["payment_stats"]["exclude"])
        )
        if payment_condition:
            self._update_detection(file_path, "payment_stats", detected)
        
        # 团购表检测
        group_condition = any(
            re.search(p, filename) 
            for p in rules["group_purchase"]["patterns"]
        )
        if group_condition:
            self._update_detection(file_path, "group_purchase", detected)

    def _update_detection(self, file_path, file_type, detected):
        """更新检测结果"""
        mtime = os.path.getmtime(file_path)
        if mtime > detected[file_type]["mtime"]:
            detected[file_type].update(path=file_path, mtime=mtime)

    def _show_detection_result(self, detected, rules):
        """显示检测结果"""
        missing = []
        print("\n" + "="*60)
        for file_type, info in detected.items():
            if info["path"]:
                mtime = datetime.datetime.fromtimestamp(info["mtime"])
                print(f"✅ {rules[file_type]['name']}：")
                print(f"   📄 文件名：{os.path.basename(info['path'])}")
                print(f"   ⏰ 修改时间：{mtime.strftime('%Y-%m-%d %H:%M:%S')}\n")
            else:
                missing.append(rules[file_type]["name"])
        
        if missing:
            print("❌ 缺失文件：")
            for name in missing:
                print(f"   ▸ {name}")
            print("\n📝 文件命名要求：")
            print("   1. 支付统计表：包含「支付方式收款统计」且不含「_已处理」")
            print("   2. 团购表：以日期开头（如：2024-03-15_团购数据.xlsx）")
        print("="*60)
        return missing

    # endregion

    # region ################### 核心处理逻辑 ###################
    def process_files(self, payment_file, group_file):
        """统一入口处理文件"""
        try:
            target_date = datetime.date.today()
            print(f"\n📅 目标处理日期：{target_date.strftime('%Y-%m-%d')}")

            group_amount = self._process_group_purchase(group_file, target_date)
            payment_data = self._process_payment_stats(payment_file)
            inputs = self._collect_user_inputs()

            self._calculate_and_show(
                payment_data,
                group_amount,
                inputs
            )
        except Exception as e:
            print(f"❌ 处理失败：{str(e)}")

    def _process_group_purchase(self, file_path, target_date):
        """处理团购表数据"""
        try:
            wb = load_workbook(file_path, data_only=True)
            total = 0.0
            date_formats = [
                # 新增支持带下划线的时间格式
                '%Y-%m-%d %H_%M_%S',
                '%Y/%m/%d %H_%M_%S',
                '%Y-%m-%d %H:%M:%S',
                '%Y/%m/%d %H:%M:%S',
                '%Y-%m-%d',
                '%Y/%m/%d',
                '%Y%m%d'
            ]

            for row in wb.active.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue

                # 日期解析增强：处理带下划线的日期格式
                cell_date = None
                raw_value = str(row[0]).strip()
                
                # 尝试提取日期部分（处理可能附加的文字）
                date_part = re.search(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', raw_value)
                if date_part:
                    raw_value = date_part.group()

                for fmt in date_formats:
                    try:
                        cell_date = datetime.datetime.strptime(raw_value, fmt).date()
                        break
                    except ValueError:
                        continue
                
                if cell_date != target_date:
                    continue

                # 列索引有效性检查
                if len(row) <= 10:
                    print(f"⚠️ 行数据列数不足，忽略该行：{row}")
                    continue

                k_value = row[10]
                try:
                    # 处理Excel可能返回的空字符串或None
                    numeric_value = float(k_value) if k_value not in (None, "", " ") else 0.0
                    total += numeric_value
                except ValueError as ve:
                    print(f"⚠️ 忽略无效数值：{k_value}，错误：{str(ve)}")
                except TypeError as te:
                    print(f"⚠️ 类型错误：{k_value}，错误：{str(te)}")
            
            print(f"ℹ️ 已处理团购表，目标日期{target_date}，累计金额：{total}")
            return total
        except Exception as e:
            print(f"❌ 团购表处理错误：{str(e)}")
            return 0.0
        
    def _process_payment_stats(self, file_path):
        """处理支付统计数据"""
        data = {
            "cash": 0.0,      # 新增现金单独统计
            "wechat": 0.0,    # 微信
            "alipay": 0.0,    # 支付宝
            "eleme": 0.0,     # 饿了么
            "member_card": 0.0,
            "times_card": 0.0
        }

        try:
            wb = load_workbook(file_path, data_only=True)
            for row in wb.active.iter_rows(values_only=True):
                if not row:
                    continue
                
                payment_type = str(row[0]).strip()
                amount = row[3] if isinstance(row[3], (int, float)) else 0.0
                
                if payment_type == "现金":
                    data["cash"] += amount
                elif payment_type == "微信":
                    data["wechat"] += amount
                elif payment_type == "支付宝":
                    data["alipay"] += amount
                elif payment_type == "饿了么":
                    data["eleme"] += amount
                elif payment_type == "余额":
                    data["member_card"] += amount
                elif payment_type == "优惠券记账金额":
                    data["times_card"] += amount
            
            # 计算零售总额
            data["retail"] = data["cash"] + data["wechat"] + data["alipay"]
            return data
        except Exception as e:
            print(f"❌ 支付统计处理错误：{str(e)}")
            return data

    # endregion

    def _collect_user_inputs(self):
        """收集用户输入数据"""
        inputs = {}
        print("\n🖍️ 请输入以下数据（直接回车默认为0）：")
        fields = [
            ("storage", "储值"),
            ("times_storage", "次卡储值"),
            ("meituan", "美团"),
            ("douyin", "抖音"),
            ("cash_total", "现金合计（实际收银现金）")  # 修改提示信息
        ]

        for key, label in fields:
            while True:
                try:
                    value = input(f"{label}: ").strip()
                    inputs[key] = float(value) if value else 0.0
                    break
                except ValueError:
                    print("⚠️ 输入无效，请重新输入数字")
        return inputs

    # endregion

    # region ################### 结果展示模块 ###################
    def _calculate_and_show(self, payment_data, group_amount, inputs):
        """计算并显示最终结果"""
        total_income = (
            payment_data["retail"] +
            inputs["meituan"] +
            payment_data["eleme"] +
            group_amount +
            inputs["douyin"] +
            inputs["storage"] +
            inputs["times_storage"]
        )

        total_sales = (
            payment_data["retail"] +
            inputs["meituan"] +
            payment_data["eleme"] +
            group_amount +
            inputs["douyin"] +
            payment_data["member_card"] +
            payment_data["times_card"]
        )

        print("\n" + "="*60)
        print("📊 最终计算结果")
        print("-"*60)

        result_items = [
            ("储值", inputs["storage"]),
            ("次卡储值", inputs["times_storage"]),
            ("零售", payment_data["retail"]),
            ("微信", payment_data["wechat"]),
            ("支付宝", payment_data["alipay"]),
            ("美团", inputs["meituan"]),
            ("饿了么", payment_data["eleme"]),
            ("抖音", inputs["douyin"]),
            ("团购", group_amount),
            ("会员卡销", payment_data["member_card"]),
            ("次卡销售", payment_data["times_card"]),
            ("今日现金", payment_data["cash"]),
            ("现金合计", inputs["cash_total"])
        ]

        def format_value(value):
            if isinstance(value, float):
                # 如果数值为整数，则直接转换为整数字符串
                if value.is_integer():
                    return str(int(value))
                else:
                    # 保留两位小数，然后去除末尾多余的0和可能的'.'
                    return f"{value:.2f}".rstrip('0').rstrip('.')
            else:
                return str(value)

        for key, value in result_items:
            print(f"{key}：{format_value(value)}")

        print(f"实收：{format_value(total_income)}")
        print(f"实销：{format_value(total_sales)}")

if __name__ == "__main__":
    app = ExcelProcessorApp()
    print("\n" + "="*60)
    print("🏷️ Excel 智能处理系统 最终版")
    print("="*60)

    try:
        mode = input("请选择模式：\n1. 自动检测文件\n2. 手动指定文件\n请输入选择：").strip()
        
        if mode == '1':
            payment_file, group_file = app.auto_detect_files()
        elif mode == '2':
            payment_file = app.sanitize_path(input("📤 支付统计文件路径："))
            group_file = app.sanitize_path(input("📥 团购文件路径："))
        else:
            print("⚠️ 无效选择，退出程序")
            exit()

        if all(os.path.exists(f) for f in (payment_file, group_file)):
            app.process_files(payment_file, group_file)
        else:
            print("❌ 错误：文件不存在")
    except KeyboardInterrupt:
        print("\n👋 用户中断操作")
    except Exception as e:
        print(f"‼️ 发生未预期错误：{str(e)}")
    finally:
        input("\n🚪 按回车键退出...")
